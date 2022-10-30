
from importlib.resources import path
import re
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut

#Es necesario poner el xlsx en la misma carpeta del código -> Sujeto a modificaciones
def open_workbook(path, sheet_name):
    global workbook 
    workbook = load_workbook(filename=path)
    if sheet_name in workbook.sheetnames:
        global sheet 
        sheet = workbook[sheet_name]
        print(f"The title of the Worksheet is: {sheet.title}")
        print(f"Cells that contain data: {sheet.calculate_dimension()}")
        global rows
        rows=sheet.max_row
        

#def write_workbook(r, c, valor):
#   c_try=sheet.cell(r,c)['E2']
#   c_try.value=valor
#   workbook.save("Personas con Discapacidad.xlsx")

"""
Los siguientes get y set son posibles porque están personalizados para este archivo xlsx, antes de usarlo con otro archivo
de diferente formato, es necesario modificarlos. Por ejemplo, en este archivo la columna 3 corresponderá siempre al 
municipio, pero esto puede variar en otro documento.
"""

def get_municipio(r):
    celda=sheet.cell(row=r, column=3)
    global municipio
    municipio=celda.value

def get_BarrVer(r):
    celda=sheet.cell(row=r, column=5)
    global barrVer
    barrVer=celda.value

def set_x(r, valor): 
    celda_x=sheet.cell(row=r, column=7)
    celda_x.value=str(valor)
    #El nombre del archivo se cambia :)

def set_y(r, valor): 
    celda_y=sheet.cell(row=r, column=8)
    celda_y.value=str(valor)
    #El nombre del archivo se cambia :)
    
#Para la geolocalizacion vamos a usar Nominatim, que es un poco inexacto pero es gratis, a diferencia de google y ArcGis
def locator(rows):
    global geolocator
    geolocator=Nominatim(user_agent="http")
    i=2
    while i <= rows:
        cell=sheet.cell(row=i, column=7).value
        if cell == 0:
            get_municipio(i)
            get_BarrVer(i)
            ubicacion=barrVer+" "+municipio+" "+"Antioquia Colombia"
            getLoc=coordBarr(ubicacion)
            if getLoc != None:
                set_x(i, getLoc.latitude)
                set_y(i, getLoc.longitude)
            cell=sheet.cell(row=i, column=7).value
        if cell == 0:
            ubicacion=municipio+" "+"Antioquia Colombia"
            getLoc=coordSinBarr(ubicacion)
            set_x(i, getLoc.latitude)
            set_y(i, getLoc.longitude)
        i+=1
        print(i)
        
    workbook.save("Personas con Discapacidad.xlsx")

def coordBarr(ubicacion):
    try:
        return geolocator.geocode(ubicacion)
    except GeocoderTimedOut:
        return coordBarr(ubicacion)
    raise


def coordSinBarr(ubicacion):
    try:
        return geolocator.geocode(ubicacion)
    except GeocoderTimedOut:
        return coordBarr(ubicacion)
    raise


if __name__== "__main__":
    open_workbook("Personas_con_Discapacidad.xlsx", sheet_name="Sin Coordenada")
    locator(rows)
